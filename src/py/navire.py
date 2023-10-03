from PyQt5.QtWidgets import *
from PyQt5 import uic
import pandas as pd
import openpyxl


class navire(QMainWindow):
    def __init__(self, parent):
        super(navire, self).__init__(parent)
        self.parent = parent
        uic.loadUi("src/ui/Navire.ui", self)
        self.toolButton.clicked.connect(self.output_file)

    def output_file(self):
        new_navire = self.lineEdit.text()
        new_IMO = self.lineEdit_2.text()
        new_AG = self.lineEdit_3.text()
        new_COM = self.lineEdit_4.text()

        # Ouvre le fichier Excel existant
        workbook = openpyxl.load_workbook("src/doc/NAVIRE.xlsx")

        # Sélectionne la première feuille de calcul du classeur
        worksheet = workbook.active

        # Ajoute les nouvelles données à la fin de la feuille de calcul
        row_num = worksheet.max_row + 1
        worksheet.cell(row=row_num, column=1, value=new_navire)
        worksheet.cell(row=row_num, column=2, value=new_IMO)
        worksheet.cell(row=row_num, column=3, value=new_AG)
        worksheet.cell(row=row_num, column=4, value=new_COM)

        # Enregistre les modifications dans le fichier Excel existant
        workbook.save("src/doc/NAVIRE.xlsx")

        self.parent.update_navire_list()
        self.close()
        