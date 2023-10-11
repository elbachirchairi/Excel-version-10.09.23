#!/usr/bin/python3
# -*- coding: utf-8 -*-
import sys
import time
# import os
import glob
import sys
import os
from datetime import datetime
import locale
from PyQt5 import QtCore, QtWidgets, QtPrintSupport, QtGui
from PyQt5.QtCore import *
from PyQt5.QtGui import QTextCursor, QIcon, QKeySequence, QColor, QTextCharFormat, QTextDocument, QTextFormat, \
    QFontDatabase
from PyQt5.QtWidgets import *
import time
from PyQt5 import QtWidgets
# import subprocess
# from PyQt5 import uic  # added
from PyQt5.QtCore import QDate,QTime
from datetime import datetime
import datetime
from datetime import datetime
# from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import openpyxl
import pandas as pd
from PyQt5 import QtWidgets, QtGui
import src
import datetime
import numpy as np
import os
from src.py.navire import navire
from src.py.flata import flata
from PyQt5.uic import loadUi
from openpyxl.workbook import Workbook
import os
import re
import glob
from PyQt5 import QtWidgets
from PyQt5 import QtCore
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles import Alignment, Font
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog
from PyQt5.QtCore import Qt
import shutil


class MyApp(QtWidgets.QMainWindow):
    def __init__(self):
        super(MyApp, self).__init__()
        self.filepath = None
        self.dirpaths = []
        loadUi('src/ui/excel.ui', self)
        self.tableWidget.setColumnWidth(0, 70)
        self.tableWidget.setColumnWidth(1, 140)
        self.tableWidget_2.setColumnWidth(0, 70)
        self.tableWidget_2.setColumnWidth(1, 140)
        self.update_navire_list()
        #self.count_non_empty_cells()
        current_date = QDate.currentDate()
        self.dateEdit_2.setDate(current_date)
        self.dateEdit_3.setDate(current_date)
        self.comboBox.currentIndexChanged.connect(self.navir)
        self.actionAjouter_navire.triggered.connect(self.add_navire)
        self.actionFata.triggered.connect(self.flata)
        self.actionClear.triggered.connect(self.CLEAR)
        self.toolButton.clicked.connect(self.input_file)
        self.toolButton_2.clicked.connect(self.input_file1)
        self.actionSave_As.triggered.connect(self.Excel1)
        self.actionEnregestrer_sous.triggered.connect(self.Excel1)
        self.toolButton_3.clicked.connect(self.input_file2)
        self.toolButton_4.clicked.connect(self.output_file)
        self.pushButton.clicked.connect(self.addlin)
        self.pushButton_2.clicked.connect(self.romovlin)
        self.pushButton_4.clicked.connect(self.addlin1)
        self.pushButton_3.clicked.connect(self.romovlin1)
        self.tableWidget.itemChanged.connect(self.count_non_empty_cells)
        self.tableWidget_2.itemChanged.connect(self.count_non_empty_cells1)

    def count_non_empty_cells1(self):
        # Compter le nombre de cellules non vides dans la colonne 0
        count = 0
        ENGINS =0
        RBN = 0
        SDFM = 0
        COL = 0
        total = 0
        for row in range(self.tableWidget_2.rowCount()):
            item = self.tableWidget_2.item(row, 0)
            item1 = self.tableWidget_2.item(row, 2)
            if item is not None and item.text().strip() != "" :  # Vérifier si l'élément n'est pas None et n'est pas vide
                count += 1  # Vérifier si la cellule n'est pas vide
                try:
                    value = int(item.text())  # Convertissez la valeur en entier
                    if 200 < value < 300:
                        RBN += 1
                    elif 100 < value < 200:
                        SDFM += 1
                except ValueError:
                    # Gérer l'erreur si la valeur de la cellule n'est pas un entier valide
                    pass
            if (item is not None and item.text().strip() != "") and (item1 is not None and item1.text().strip() != ""):  # Vérifier si l'élément n'est pas None et n'est pas vide
                count -= 1 
            if (item is not None and item.text().strip() == "") and (item1 is not None and item1.text().strip() != ""):  # Vérifier si l'élément n'est pas None et n'est pas vide
                count += 1 
        total = count - COL
        self.lineEdit_11.setText(str(total))
        self.lineEdit_12.setText(str(SDFM))
        self.lineEdit_13.setText(str(RBN))
        self.lineEdit_9.setText(str(COL))
        self.lineEdit_8.setText(str(ENGINS))
    
    def count_non_empty_cells(self):
        # Compter le nombre de cellules non vides dans la colonne 0
        count = 0
        ENGINS =0
        RBN = 0
        SDFM = 0
        COL = 0
        total = 0
        for row in range(self.tableWidget.rowCount()):
            item = self.tableWidget.item(row, 0)
            item1 = self.tableWidget.item(row, 2)
            item3 = self.tableWidget.item(row, 1)
            if (item is not None and item.text().strip() != "") or (item1 is not None and item1.text().strip() != ""):
                # Au moins l'un des éléments n'est pas vide
                count += 1

                try:
                    value = int(item.text())  # Convertissez la valeur en entier
                    if 200 < value < 300:
                        RBN += 1
                    elif 100 < value < 200:
                        SDFM += 1

                except ValueError:
                    # Gérer l'erreur si la valeur de la cellule n'est pas un entier valide
                    pass
            elif item1 is not None and item1.text().strip() != "":
                count += 1


        total = count - COL
        self.lineEdit_4.setText(str(total))
        self.lineEdit_2.setText(str(SDFM))
        self.lineEdit_3.setText(str(RBN))
        self.lineEdit_5.setText(str(COL))
        self.lineEdit_6.setText(str(ENGINS))
    def romovlin1(self):
        selected_row = self.tableWidget_2.currentRow()  # Obtenez le numéro de la ligne sélectionnée
        if selected_row >= 0:
            self.tableWidget_2.removeRow(selected_row)
        self.count_non_empty_cells1()
    
    def addlin1(self):
        row = self.tableWidget_2.rowCount()
        self.tableWidget_2.insertRow(row)

    def romovlin(self):
        selected_row = self.tableWidget.currentRow()  # Obtenez le numéro de la ligne sélectionnée
        if selected_row >= 0:
            self.tableWidget.removeRow(selected_row)
        self.count_non_empty_cells()
    
    def addlin(self):
        row = self.tableWidget.rowCount()
        self.tableWidget.insertRow(row)
    
    def output_file(self):
        import openpyxl

        file = str(self.lineEdit_15.text())
        if file == "":
            pass
        else:
            locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            date = self.dateEdit_3.date().toPyDate()
            mois_fr = date.strftime('%B').upper()
            navire = self.comboBox.currentText()
            date_str = date.strftime('%d/%m/%Y')
            ED = str(self.lineEdit_16.text())
            EMB = str(self.lineEdit_17.text())
            srimport = int(self.lineEdit_4.text())
            collimport = int(self.lineEdit_5.text())
            srexport = int(self.lineEdit_11.text())
            collexport = int(self.lineEdit_9.text())
            rta = self.timeEdit_5.time().toPyTime()
            rta_str = rta.strftime("%HH%M")
            rtd = self.timeEdit_11.time().toPyTime()
            rtd_str = rtd.strftime("%HH%M")
            TOTALIMPORT = srimport + collimport
            TOTALEXPORT = srexport + collexport
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook[mois_fr]
            bold_font = Font(bold=True)
            border_style = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
            row_num = worksheet.max_row + 1
            worksheet.cell(row=row_num, column=1, value=date_str)
            if TOTALIMPORT == 0:
                worksheet.cell(row=row_num, column=2, value="0")
                worksheet.cell(row=row_num, column=7, value="0")
                worksheet.cell(row=row_num, column=8, value="0")
                worksheet.cell(row=row_num, column=13, value="0")
                worksheet.cell(row=row_num, column=14, value="0")
            else:
                worksheet.cell(row=row_num, column=2, value=ED)
                worksheet.cell(row=row_num, column=7, value="OK")
                worksheet.cell(row=row_num, column=8, value="OK")
                worksheet.cell(row=row_num, column=13, value=TOTALIMPORT)
                worksheet.cell(row=row_num, column=14, value=TOTALIMPORT)

            if TOTALEXPORT == 0:
                worksheet.cell(row=row_num, column=3, value="0")
                worksheet.cell(row=row_num, column=9, value="0")
                worksheet.cell(row=row_num, column=10, value="0")
                worksheet.cell(row=row_num, column=11, value="0")
                worksheet.cell(row=row_num, column=12, value="0")
                worksheet.cell(row=row_num, column=15, value="0")
                worksheet.cell(row=row_num, column=16, value="0")
            else:
                worksheet.cell(row=row_num, column=3, value=EMB)
                worksheet.cell(row=row_num, column=9, value="OK")
                worksheet.cell(row=row_num, column=10, value="OK")
                worksheet.cell(row=row_num, column=11, value="OK")
                worksheet.cell(row=row_num, column=12, value="OK")
                worksheet.cell(row=row_num, column=15, value=TOTALEXPORT)
                worksheet.cell(row=row_num, column=16, value=TOTALEXPORT)
            worksheet.cell(row=row_num, column=4, value=navire)
            worksheet.cell(row=row_num, column=5, value=rta_str)
            worksheet.cell(row=row_num, column=6, value=rtd_str)
            worksheet.cell(row=row_num, column=17, value="RIEN")
            worksheet.cell(row=row_num, column=18, value="RIEN")
            worksheet.cell(row=row_num, column=19, value="RIEN")
            worksheet.cell(row=row_num, column=23, value="RIEN")
            worksheet.cell(row=row_num, column=27, value="RIEN")
            worksheet.merge_cells(start_row=row_num, start_column=19, end_row=row_num, end_column=22)
            worksheet.merge_cells(start_row=row_num, start_column=23, end_row=row_num, end_column=26)

            for row in worksheet.rows:
                for cell in row:
                    if cell.value:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:  # vérifie si la cellule est remplie
                        cell.border = border_style
            # border_bottom = Border(bottom=Side(border_style='thin', color='000000'))
            for cell in worksheet[f"S{row_num}:V{row_num}"][0]:
                cell.border = border_style

            for cell in worksheet[f"W{row_num}:AA{row_num}"][0]:
                cell.border = border_style
            workbook.save(file)

    def input_file2(self):
        self.filepath = QtWidgets.QFileDialog.getOpenFileName(self, "Select excel File", "~","*.xlsx;;All Files(*)")[0]
        if self.filepath == "":
            pass
        else:
            self.lineEdit_15.setText(str(self.filepath))
    def Excel1(self):
        import openpyxl
        if self.timeEdit_5.time()== QTime(0, 0):
            QMessageBox.warning(self, "Warning", "Vérifier RTA", QMessageBox.Ok)
        if self.timeEdit_11.time()== QTime(0, 0):
            QMessageBox.warning(self, "Warning", "Vérifier RTD", QMessageBox.Ok)
        if self.timeEdit_4.time()== QTime(0, 0):
            QMessageBox.warning(self, "Warning", "Vérifier ETA", QMessageBox.Ok)
        if self.timeEdit_10.time()== QTime(0, 0):
            QMessageBox.warning(self, "Warning", "Vérifier ETD", QMessageBox.Ok)
        if self.comboBox.currentIndex() == 0:
            QMessageBox.warning(self, "Warning", "Choisissez un navire", QMessageBox.Ok)
            return
        elif int(self.lineEdit_4.text())>0 and self.lineEdit_16.text()=="":
            QMessageBox.warning(self, "Warning", "Donnez le n° d'état de déchargement.", QMessageBox.Ok)
            return
        elif int(self.lineEdit_11.text())>0 and self.lineEdit_17.text()=="":
            QMessageBox.warning(self, "Warning", "Donnez le n° d'embarquement.", QMessageBox.Ok)
            return
        else :
            workbook = openpyxl.load_workbook('src/doc/roro1.xlsx')
            #---------------------------------------------------------
            #************************* IMPORT **************************
            sheet = workbook['IMPORT']
            date = self.dateEdit_2.date().toPyDate()
            date_str = date.strftime('%d/%m/%Y')
            navire = str(self.comboBox.currentText())
            eta = self.timeEdit_4.time().toPyTime()
            eta_str = eta.strftime("%HH%M")
            rta = self.timeEdit_5.time().toPyTime()
            rta_str = rta.strftime("%HH%M")
            db = self.timeEdit_6.time().toPyTime()
            db_str = db.strftime("%HH%M")
            fb = self.timeEdit_7.time().toPyTime()
            fb_str = fb.strftime("%HH%M")
            ed = self.lineEdit_16.text()
            total = self.lineEdit_4.text()
            sdfm = self.lineEdit_2.text()
            rbn = self.lineEdit_3.text()
            cal = self.lineEdit_5.text()
            eng = self.lineEdit_6.text()
            pointeur = str(self.lineEdit.text())
            sheet["A4"] = pointeur
            sheet["A5"] = "DATE: " + str(date_str)
            sheet["A6"] = "NOM DU NAVIRE: " + navire
            sheet["D5"] = "ETA:  " + eta_str
            sheet["D6"] = "RTA:  " + rta_str
            sheet["A7"] = "HEURE DEBUT DE DECHARGEMENT:  " + db_str
            sheet["D7"] = "HEURE FIN DE DECHARGEMENT: " + fb_str
            sheet["C10"] = "N° D'ETAT DE DECHARGEMENT: "+ed
            sheet["D9"] = total
            sheet["C8"] = sdfm
            sheet["C9"] = rbn
            sheet["E9"] = cal
            sheet["F9"] = eng

            tableWidget = self.tableWidget
            rowCount = tableWidget.rowCount()
            columnCount = tableWidget.columnCount()
            start_row = 13
            start_column = 2
            i = 1
            for row in range(rowCount):
                for col in range(columnCount):
                    cell_text = str(tableWidget.item(row, col).text())

                    if not cell_text.strip():
                        cell_text = '=" "'

                    sheet.cell(row=start_row + row, column=start_column + col, value=cell_text)
                sheet.cell(row=start_row + row, column=1, value=i)
                i += 1

            #---------------------------------------------------------
            #********************** EXPORT ************************
            sheet = workbook['EXPORT']
            locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            date1 = self.dateEdit_3.date().toPyDate()
            mo_fr = date1.strftime('%B').upper()
            anneetext = str(date1.year)
            date1_str = date1.strftime('%d/%m/%Y')
            date2_str = date1.strftime('%d-%m-%Y')
            navire1 = str(self.comboBox_2.currentText())
            etd = self.timeEdit_8.time().toPyTime()
            etd_str = etd.strftime("%HH%M")
            rtd = self.timeEdit_11.time().toPyTime()
            rtd_str = rtd.strftime("%HH%M")
            db1 = self.timeEdit_9.time().toPyTime()
            db1_str = db1.strftime("%HH%M")
            fb1 = self.timeEdit_10.time().toPyTime()
            fb1_str = fb1.strftime("%HH%M")
            em = self.lineEdit_17.text()
            total1 = self.lineEdit_11.text()
            sdfm1 = self.lineEdit_12.text()
            rbn1 = self.lineEdit_13.text()
            cal1 = self.lineEdit_9.text()
            eng1 = self.lineEdit_8.text()
            pointeur1 = str(self.lineEdit_10.text())
            sheet["A5"] = "DATE: " + str(date1_str)
            sheet["A4"] = pointeur1
            sheet["A6"] = "NOM DU NAVIRE: " + navire1
            sheet["D5"] = "ETD:  " + etd_str
            sheet["D6"] = "RTD:  " + rtd_str
            sheet["A7"] = "HEURE DEBUT DE CHARGEMENT:   " + db1_str
            sheet["D7"] = "HEURE FIN DE CHARGEMENT : " + fb1_str
            sheet["C10"] = "N° D'EMBARQUEMENT: "+em
            sheet["D9"] = total1
            sheet["C8"] = sdfm1
            sheet["C9"] = rbn1
            sheet["E9"] = cal1
            sheet["F9"] = eng1
            tableWidget_2 = self.tableWidget_2
            rowCount = tableWidget_2.rowCount()
            columnCount = tableWidget_2.columnCount()
            start_row = 13
            start_column = 2
            i = 1
            for row in range(rowCount):
                for col in range(columnCount):
                    cell_text = str(tableWidget_2.item(row, col).text())

                    if not cell_text.strip():
                        cell_text = '=" "'

                    sheet.cell(row=start_row + row, column=start_column + col, value=cell_text)
                sheet.cell(row=start_row + row, column=1, value=i)
                i += 1
            self.filepath = QtWidgets.QFileDialog.getExistingDirectory(self, "Save as", "")
            filename = str(navire +" " + date2_str + " à " + etd_str + ".xlsx")
            save_file = self.filepath + "/" + filename
            workbook.save(save_file)
            destination_path = "src/Archive/Excels/" + anneetext + "/" + str(mo_fr) +"/" + filename
            shutil.copy(save_file, destination_path)
            #//////////////////////////////////////////////////////////
            locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
            date = self.dateEdit_3.date().toPyDate()

            date_RTA = self.dateEdit_2.date().toPyDate()
            annee_text = str(date_RTA.year)
            mois_fr = date_RTA.strftime('%B').upper()
            navire = self.comboBox.currentText()
            date_str = date.strftime('%d/%m/%Y')
            date_strRTA = date_RTA.strftime('%d/%m/%Y')
            ED = str(self.lineEdit_16.text())
            EMB = str(self.lineEdit_17.text())
            srimport = int(self.lineEdit_4.text())
            collimport = int(self.lineEdit_5.text())
            srexport = int(self.lineEdit_11.text())
            collexport = int(self.lineEdit_9.text())
            rta = self.timeEdit_5.time().toPyTime()
            rta_str = rta.strftime("%HH%M")
            rtd = self.timeEdit_11.time().toPyTime()
            rtd_str = rtd.strftime("%HH%M")
            TOTALIMPORT = srimport + collimport
            TOTALEXPORT = srexport + collexport
            workbook = openpyxl.load_workbook("src/Archive/Excels/"+ annee_text + "/Suivi des navires RORO " + annee_text + ".xlsx")
            worksheet = workbook[mois_fr]
            bold_font = Font(bold=True)
            border_style = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
            row_num = worksheet.max_row + 1
            worksheet.cell(row=row_num, column=1, value=date_strRTA)
            worksheet.cell(row=row_num, column=2, value=date_str)
            if TOTALIMPORT == 0:
                worksheet.cell(row=row_num, column=3, value="0")
                worksheet.cell(row=row_num, column=8, value="0")
                worksheet.cell(row=row_num, column=9, value="0")
                worksheet.cell(row=row_num, column=14, value="0")
                worksheet.cell(row=row_num, column=15, value="0")
            else:
                worksheet.cell(row=row_num, column=3, value=ED)
                worksheet.cell(row=row_num, column=8, value="OK")
                worksheet.cell(row=row_num, column=9, value="OK")
                worksheet.cell(row=row_num, column=14, value=TOTALIMPORT)
                worksheet.cell(row=row_num, column=15, value=TOTALIMPORT)

            if TOTALEXPORT == 0:
                worksheet.cell(row=row_num, column=4, value="0")
                worksheet.cell(row=row_num, column=10, value="0")
                worksheet.cell(row=row_num, column=11, value="0")
                worksheet.cell(row=row_num, column=12, value="0")
                worksheet.cell(row=row_num, column=13, value="0")
                worksheet.cell(row=row_num, column=16, value="0")
                worksheet.cell(row=row_num, column=17, value="0")
            else:
                worksheet.cell(row=row_num, column=4, value=EMB)
                worksheet.cell(row=row_num, column=10, value="OK")
                worksheet.cell(row=row_num, column=11, value="OK")
                worksheet.cell(row=row_num, column=12, value="OK")
                worksheet.cell(row=row_num, column=13, value="OK")
                worksheet.cell(row=row_num, column=16, value=TOTALEXPORT)
                worksheet.cell(row=row_num, column=17, value=TOTALEXPORT)
            worksheet.cell(row=row_num, column=5, value=navire)
            worksheet.cell(row=row_num, column=6, value=rta_str)
            worksheet.cell(row=row_num, column=7, value=rtd_str)
            worksheet.cell(row=row_num, column=18, value="RIEN")
            worksheet.cell(row=row_num, column=19, value="RIEN")
            worksheet.cell(row=row_num, column=20, value="RIEN")
            worksheet.cell(row=row_num, column=24, value="RIEN")
            worksheet.cell(row=row_num, column=28, value="RIEN")
            worksheet.merge_cells(start_row=row_num, start_column=20, end_row=row_num, end_column=23)
            worksheet.merge_cells(start_row=row_num, start_column=24, end_row=row_num, end_column=27)

            for row in worksheet.rows:
                for cell in row:
                    if cell.value:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:  # vérifie si la cellule est remplie
                        cell.border = border_style
            # border_bottom = Border(bottom=Side(border_style='thin', color='000000'))
            for cell in worksheet[f"S{row_num}:V{row_num}"][0]:
                cell.border = border_style

            for cell in worksheet[f"W{row_num}:AA{row_num}"][0]:
                cell.border = border_style
            workbook.save("src/Archive/Excels/"+ annee_text + "/Suivi des navires RORO " + annee_text + ".xlsx")
            #///////////////////////////
            workbook = openpyxl.load_workbook("src/Archive/Décade/"+ annee_text + "/" + str(mois_fr) + "/Décade.xlsx")
            worksheet = workbook.active
            bold_font = Font(bold=True)
            border_style = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None or cell.value == 0:
                        cell.border = border_style
            row_num = worksheet.max_row + 1
            worksheet.cell(row=row_num, column=2, value=navire)
            worksheet.cell(row=row_num+1, column=2, value=navire)
            worksheet.cell(row=row_num, column=3, value=ED)
            worksheet.cell(row=row_num+1, column=3, value=EMB)
            worksheet.cell(row=row_num, column=4, value=date_strRTA)
            worksheet.cell(row=row_num+1, column=4, value=date_str)
            worksheet.cell(row=row_num, column=5, value=rta_str)
            worksheet.cell(row=row_num + 1, column=5, value=rtd_str)
            worksheet.cell(row=row_num, column=6, value=srimport)
            worksheet.cell(row=row_num+1, column=6, value=srexport)
            worksheet.cell(row=row_num, column=7, value=collimport)
            worksheet.cell(row=row_num+1, column=7, value=collexport)



            for row in worksheet.rows:
                for cell in row:
                    if cell.value:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:  # vérifie si la cellule est remplie
                        cell.border = border_style
            # border_bottom = Border(bottom=Side(border_style='thin', color='000000'))
            for cell in worksheet[f"S{row_num}:V{row_num}"][0]:
                cell.border = border_style

            for cell in worksheet[f"W{row_num}:AA{row_num}"][0]:
                cell.border = border_style

            workbook.save("src/Archive/Décade/"+ annee_text + "/" + str(mois_fr) + "/Décade.xlsx")


    def input_file1(self):
        self.filepath = QtWidgets.QFileDialog.getOpenFileName(self, "Select excel File", "~", "*.xlsx;;All Files(*)")[0]
        if self.filepath == "":
            pass
        else:
            MAFI = []
            Unit = []
            Rectifier = []
            self.lineEdit_14.setText(str(self.filepath))
            df = pd.read_excel(self.filepath)
            if "MAFI" in df.columns or "Unité" in df.columns or "???fr:Rectifier???" in df.columns:
                MAFI = df["MAFI"].tolist()
                Unit = df["Unité"].tolist()
                Rectifier = df["???fr:Rectifier???"].tolist()
            nombre_de_valeurs_mafi = len(MAFI)
            for i in range(nombre_de_valeurs_mafi):
                if pd.isna(Rectifier[i]):
                    Rectifier[i] = " "
                if pd.isna(MAFI[i]):
                    MAFI[i] = ""
                if MAFI[i] == "":
                    # Si MAFI[i] est vide, passez à l'itération suivante de la boucle
                    continue
                if MAFI[i] == 0:
                    MAFI[i] = "0"
                    Rectifier[i] = "SANS MAFI"
                row = self.tableWidget_2.rowCount()
                self.tableWidget_2.insertRow(row)
                item_mafi = QTableWidgetItem(str(int(MAFI[i])))
                item_mafi.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_2.setItem(row, 0, item_mafi)
                item_unit = QTableWidgetItem(str(Unit[i]))
                item_unit.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_2.setItem(row, 1, item_unit)
                item_rectifier = QTableWidgetItem(str(Rectifier[i]))
                item_rectifier.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_2.setItem(row, 2, item_rectifier)

        ENGINS = 0
        RBN = 0
        SDFM = 0
        COL = 0
        total = 0

        # Parcourez les lignes de la première colonne du QTableWidget
        for i in range(self.tableWidget_2.rowCount()):
            item = self.tableWidget_2.item(i, 0)  # Obtenez l'élément de la première colonne
            if item is not None and not item.text().strip():  # Vérifiez si la cellule est vide
                continue

            value = int(item.text())  # Convertissez la valeur en entier
            if 200 < value < 300:
                RBN += 1
            elif 100 < value < 200:
                SDFM += 1

            total = i + 1 - COL
        self.lineEdit_11.setText(str(total))
        self.lineEdit_12.setText(str(SDFM))
        self.lineEdit_13.setText(str(RBN))
        self.lineEdit_9.setText(str(COL))
        self.lineEdit_8.setText(str(ENGINS))

    def input_file(self):
        self.filepath = QtWidgets.QFileDialog.getOpenFileName(self, "Select excel File", "~", "*.xlsx;;All Files(*)")[0]
        if self.filepath == "":
            pass
        else:
            MAFI = []
            Unit = []
            Rectifier = []
            self.lineEdit_7.setText(str(self.filepath))
            df = pd.read_excel(self.filepath)
            if "MAFI" in df.columns or "Unité" in df.columns or "???fr:Rectifier???" in df.columns:
                MAFI = df["MAFI"].tolist()
                Unit = df["Unité"].tolist()
                Rectifier = df["???fr:Rectifier???"].tolist()
            nombre_de_valeurs_mafi = len(MAFI)
            for i in range(nombre_de_valeurs_mafi):
                if pd.isna(Rectifier[i]):
                    Rectifier[i] = " "
                if pd.isna(MAFI[i]):
                    MAFI[i] = ""
                if MAFI[i] == "":
                    # Si MAFI[i] est vide, passez à l'itération suivante de la boucle
                    continue
                if MAFI[i] == 0:
                    MAFI[i] = "0"
                    Rectifier[i] = "SANS MAFI"
                row = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row)
                item_mafi = QTableWidgetItem(str(int(MAFI[i])))
                item_mafi.setTextAlignment(Qt.AlignCenter)
                self.tableWidget.setItem(row, 0, item_mafi)
                item_unit = QTableWidgetItem(str(Unit[i]))
                item_unit.setTextAlignment(Qt.AlignCenter)
                self.tableWidget.setItem(row, 1, item_unit)
                item_rectifier = QTableWidgetItem(str(Rectifier[i]))
                item_rectifier.setTextAlignment(Qt.AlignCenter)
                self.tableWidget.setItem(row, 2, item_rectifier)

        ENGINS = 0
        RBN = 0
        SDFM = 0
        COL = 0
        total = 0

        # Parcourez les lignes de la première colonne du QTableWidget
        for i in range(self.tableWidget.rowCount()):
            item = self.tableWidget.item(i, 0)  # Obtenez l'élément de la première colonne
            if item is not None and not item.text().strip():  # Vérifiez si la cellule est vide
                continue

            value = int(item.text())  # Convertissez la valeur en entier
            if 200 < value < 300:
                RBN += 1
            elif 100 < value < 200:
                SDFM += 1

            total = i + 1 - COL
        self.lineEdit_4.setText(str(total))
        self.lineEdit_2.setText(str(SDFM))
        self.lineEdit_3.setText(str(RBN))
        self.lineEdit_5.setText(str(COL))
        self.lineEdit_6.setText(str(ENGINS))


    def CLEAR(self):
        self.comboBox.setCurrentIndex(0)
        self.tableWidget.setRowCount(0)
        self.tableWidget_2.setRowCount(0)
        #self.lineEdit_2.clear()
        self.lineEdit_2.setText("0")
        self.lineEdit_3.setText("0")
        self.lineEdit_4.setText("0")
        self.lineEdit_5.setText("0")
        self.lineEdit_6.setText("0")
        self.lineEdit_8.setText("0")
        self.lineEdit_9.setText("0")
        self.lineEdit_11.setText("0")
        self.lineEdit_12.setText("0")
        self.lineEdit_13.setText("0")
        self.lineEdit_16.setText("")
        self.lineEdit_17.setText("")
        self.lineEdit_7.setText("")
        self.lineEdit_14.setText("")
        self.lineEdit_15.setText("")
        self.lineEdit_10.setText(" Nom & Prénom de l'agent pointeur : ")
        self.lineEdit.setText(" Nom & Prénom de l'agent pointeur : ")
        current_date = QDate.currentDate()
        self.dateEdit_2.setDate(current_date)
        self.dateEdit_3.setDate(current_date)
        #/////////////////
        self.timeEdit_4.setDisplayFormat("'ETA' :                      HH'H'mm")
        time = self.timeEdit_4.time()
        time.setHMS(0, 0, 0)
        self.timeEdit_4.setTime(time)
        self.timeEdit_4.setDisplayFormat("'RTA :'                      HH'H'mm")
        time1 = self.timeEdit_5.time()
        time1.setHMS(0, 0, 0)
        self.timeEdit_5.setTime(time1)
        self.timeEdit_6.setDisplayFormat("'HEURE DEBUT DE DECHARGEMENT: 'HH'H'mm")
        time2 = self.timeEdit_6.time()
        time2.setHMS(0, 0, 0)
        self.timeEdit_6.setTime(time2)
        self.timeEdit_7.setDisplayFormat("'HEURE FIN DE DECHARGEMENT:' HH'H'mm")
        time3 = self.timeEdit_7.time()
        time3.setHMS(0, 0, 0)
        self.timeEdit_7.setTime(time3)
        # /////////////////
        self.timeEdit_8.setDisplayFormat("'ETD' :                      HH'H'mm")
        time4 = self.timeEdit_8.time()
        time4.setHMS(0, 0, 0)
        self.timeEdit_8.setTime(time4)
        self.timeEdit_11.setDisplayFormat("'RTD :'                      HH'H'mm")
        time5 = self.timeEdit_11.time()
        time5.setHMS(0, 0, 0)
        self.timeEdit_11.setTime(time5)
        self.timeEdit_9.setDisplayFormat("'HEURE DEBUT DE CHARGEMENT: 'HH'H'mm")
        time6 = self.timeEdit_9.time()
        time6.setHMS(0, 0, 0)
        self.timeEdit_9.setTime(time6)
        self.timeEdit_10.setDisplayFormat("'HEURE FIN DE CHARGEMENT :' HH'H'mm")
        time7 = self.timeEdit_10.time()
        time7.setHMS(0, 0, 0)
        self.timeEdit_10.setTime(time7)

    def update_navire_list(self):
        # mise à jour de la comboBox avec les navires existants
        self.comboBox.clear()
        item1 = QIcon("src/img/navire.png")
        data = pd.read_excel("src/doc/NAVIRE.xlsx")
        navire_list = data['NAVIRE'].tolist()
        for i, navire in enumerate(navire_list):
            self.comboBox.addItem(str(navire))
            if i == 0:
                self.comboBox.setItemData(i, QIcon("src/img/navire.png"), QtCore.Qt.DecorationRole)

    def add_navire(self):
        self.Navire = navire(parent=self)
        # self.Navire.accepted.connect(self.update_navire_list)
        self.Navire.show()

    def flata(self):
        self.flata = flata(parent=self)
        # self.Navire.accepted.connect(self.update_navire_list)
        self.flata.show()

    def navir(self):
        navire = self.comboBox.currentText()

        # Charger les données de l'Excel
        data = pd.read_excel("src/doc/NAVIRE.xlsx", engine='openpyxl')

        # Vérifier si la valeur sélectionnée est présente dans le dataframe
        if navire in data['NAVIRE'].values:
            # Rechercher le nom de la compagnie maritime correspondant au navire sélectionné
            compagnie_maritime = data.loc[data['NAVIRE'] == navire, 'COMPAGNIE MARITIME'].iloc[0]
            agent_maritime = data.loc[data['NAVIRE'] == navire, 'AGENT MARITIME'].iloc[0]
            # Afficher le nom de la compagnie maritime dans un label
            self.lineEdit_18.setText(str(compagnie_maritime))
            self.lineEdit_19.setText(str(agent_maritime))
            self.lineEdit_21.setText(str(compagnie_maritime))
            self.lineEdit_20.setText(str(agent_maritime))
            currentText = str(self.comboBox.currentText())
            self.comboBox_2.clear()
            self.comboBox_2.addItem(currentText)
            if self.comboBox.currentIndex() == 0:
                self.comboBox_2.setItemIcon(0, QIcon("src/img/navire.png"))

        else:
            # Gérer le cas où la valeur sélectionnée n'est pas présente dans le dataframe
            self.lineEdit_18.setText('')
            self.lineEdit_19.setText('')
            self.lineEdit_20.setText('')
            self.lineEdit_21.setText('')
            self.comboBox_2.clear()
            self.comboBox_2.addItem('')
            QMessageBox.warning(self, 'Attention', 'La valeur sélectionnée n\'est pas présente dans le dataframe.')


if __name__ == '__main__':
    my_app = QtWidgets.QApplication(sys.argv)
    my_window = MyApp()
    my_window.show()
    my_window.setStyleSheet(
        # "QPushButton { background-color: palegoldenrod; border-width: 2px; border-color: darkkhaki}"
        # "QPushButton { border-style: solid; border-radius: 5; padding: 3px; min-width: 9ex; min-height: 2.5ex;}"
        # "QLabel, QAbstractButton { font: bold; font-size: 9px }"
        "QPushButton#evilButton { background-color: palegoldenrod;border-style: outset;border-width: 2px;border-radius: 10px;border-color: darkkhaki ;font: bold 14px;min-width: 10em;padding: 6px}"
        "QToolButton#evilButton { background-color: palegoldenrod;border-style: outset;border-width: 2px;border-radius: 10px;border-color: darkkhaki ;font: bold 14px;min-width: 10em;padding: 6px}"
        #"QStatusBar QLabel { font: normal }"
        "QStatusBar::item { border-width: 1; border-color: darkkhaki; border-style: solid; border-radius: 2;}"
        "QLineEdit, QSpinBox, QTextEdit, QListView { background-color: cornsilk; selection - color: #0a214c}"
        " QLineEdit, QSpinBox, QTextEdit, QListView { selection-background-color:  #C19A6B;}"
        "QLineEdit, QFrame { border-width: 1px; border-style: solid; border-color: darkkhaki; border-radius: 5px}"
        "QLabel { border: none; padding: 0; background: none; }"
        "QMenuBar {background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,stop:0 lightgray, stop:1 #FFFDE2);spacing: 3px; /* spacing between menu bar items */}"
        "QPlainTextEdit {font-family: Monospace; font-size: 13px; background:  #E2E2E2; color:  #202020; border: 1px solid #1EAE3D;}")
    sys.exit(my_app.exec())